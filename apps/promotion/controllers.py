from flask import render_template, jsonify, abort, request, send_file, current_app
import re
from . import promote
from flask_login import current_user, login_required
from ..decorators import student_required, parent_required, school_it_admin_required
from ..models.school.student import Student
from ..models.general.company import Company
from ..models.school.session import Session
from ..models.school.subject import Subject
from ..models.school.academic_year import AcademicYear
from ..models.school.exam import Exam
from ..models.school.grade import Grade
from ..models.general.user import User
from ..models.school.classroom import Class
from ..models.school.parent import Parent
from io import BytesIO
import qrcode
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from .utils import generate_qr_code, generate_report_token
from flask_babel import gettext as _
from .utils import compute_student_report_fr, get_previous_sessions, calculate_annual_gpa_and_rank
from weasyprint import HTML
from werkzeug.utils import secure_filename
import zipfile
from .utils import ReportToken
import os

"""
school related
system
"""
"""route for students to view their own grades"""
@promote.route('/student_grades/<int:company_id>', methods=['GET'])
@login_required
@student_required
def view_my_grades(company_id):
    company = Company.query.get_or_404(company_id)
    
    # query the student info
    student = User.query.join(User.students).filter(Student.user_id == current_user.id).first_or_404()
    
    # Retrieve the student's class
    student_class = Class.query.join(Class.students).filter(Student.user_id == student.id).first_or_404()

    subject_id = request.args.get('subject_id')
    session_id = request.args.get('session_id')
    exam_id = request.args.get('exam_id')
    academic_year_id = request.args.get('academic_year_id')

    # Get the subjects, sessions, exams, and academic years relevant to the student
    subjects = Subject.query.join(Student.subjects).filter(Student.user_id == student.id).all()
    
    sessions = Session.query.join(AcademicYear).filter(
        Session.company_id == company.id,
        AcademicYear.active == True
    ).all()

    # Get the exams based on the class the student is in
    exams = Exam.query.filter_by(class_id=student_class.id).all()

    academic_years = AcademicYear.query.filter_by(company_id=company.id).all()

    # Build the grades query with the selected filters
    grades_query = Grade.query.filter_by(student_id=student.id)

    if subject_id:
        grades_query = grades_query.filter_by(subject_id=subject_id)
    if session_id:
        grades_query = grades_query.filter_by(session_id=session_id)
    if exam_id:
        grades_query = grades_query.filter_by(exam_id=exam_id)
    if academic_year_id:
        grades_query = grades_query.join(Session).filter(Session.academic_year_id == academic_year_id)

    grades = grades_query.all()

    subject_dict = {subject.id: subject.name for subject in subjects}
    session_dict = {session.id: session.name for session in sessions}
    exam_dict = {exam.id: exam.name for exam in exams}

    return render_template(
        'views/school/student/grades.html',
        company=company,
        subjects=subjects,
        sessions=sessions,
        academic_years=academic_years,
        grades=grades,
        student=student,
        exams=exams,
        subject_dict=subject_dict,
        session_dict=session_dict,
        exam_dict=exam_dict
    )


"""
route for parents to view their children grades
"""
@promote.route('/parent_grades/<int:company_id>', methods=['GET'])
@login_required
@parent_required
def view_child_grades(company_id):
    company = Company.query.get_or_404(company_id)
    
    parent = Parent.query.filter_by(user_id=current_user.id).first_or_404()
    

    students = Student.query.filter_by(parent_id=parent.id).all()
    
    all_grades = {}
    all_subjects = {}
    all_sessions = {}
    all_exams = {}
    
    subject_id = request.args.get('subject_id')
    session_id = request.args.get('session_id')
    exam_id = request.args.get('exam_id')
    
    # Iterate through each child and retrieve their grades
    for student in students:
        student_class = Class.query.join(Class.students).filter(Student.user_id == student.user_id).first_or_404()
        subjects = Subject.query.join(Student.subjects).filter(Student.user_id == student.user_id).all()
        sessions = Session.query.join(AcademicYear).filter(
            Session.company_id == company.id,
            AcademicYear.active == True
        ).all()

        exams = Exam.query.filter_by(class_id=student_class.id).all()


        grades_query = Grade.query.filter_by(student_id=student.user_id)
        
        if subject_id:
            grades_query = grades_query.filter_by(subject_id=subject_id)
        if session_id:
            grades_query = grades_query.filter_by(session_id=session_id)
        if exam_id:
            grades_query = grades_query.filter_by(exam_id=exam_id)

        grades = grades_query.all()

        all_grades[student.id] = grades
        all_subjects[student.id] = subjects
        all_sessions[student.id] = sessions
        all_exams[student.id] = exams
    
    return render_template(
        'views/school/parent/grades.html',
        company=company,
        students=students,
        all_grades=all_grades,
        all_subjects=all_subjects,
        all_sessions=all_sessions,
        all_exams=all_exams
    )


@promote.route('/generate_session_reports/<int:class_id>/<int:session_id>')
@school_it_admin_required
@login_required
def generate_session_reports(class_id, session_id):
    class_ = Class.query.get_or_404(class_id)
    session = Session.query.get_or_404(session_id)

    user_downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads", "reports")
    os.makedirs(user_downloads_dir, exist_ok=True)

    pdfs_dir = os.path.join(user_downloads_dir, f'class_{class_id}_session_{session_id}')
    os.makedirs(pdfs_dir, exist_ok=True)

    for student in class_.students:
        try:
            # Compute current session report
            data = compute_student_report_fr(student.user_id, session_id)

            previous_sessions = get_previous_sessions(session.academic_year_id, session.start_date)
            annual_gpa, annual_rank = calculate_annual_gpa_and_rank(student.user_id, previous_sessions)

            # Include annual results in the report data
            data['annual_gpa'] = annual_gpa
            data['annual_rank'] = annual_rank

            if class_.grading_system == 'FrenchSec':
                html = render_template('reports/academics/FR/session_report_primary.html', data=data)
            else:
                html = render_template('reports/academics/FR/session_report_high_school.html', data=data)

            pdf_file_name = f"{secure_filename(student.user.last_name)}_{secure_filename(student.user.first_name)}_report.pdf"
            pdf_file_path = os.path.join(pdfs_dir, pdf_file_name)
            HTML(string=html).write_pdf(pdf_file_path)
        except Exception as e:
            current_app.logger.error(f"Failed to generate report for {student.user_id}: {e}")
            continue

    zip_filename = f'class_{class_id}_session_{session_id}_reports.zip'
    zip_path = os.path.join(user_downloads_dir, zip_filename)

    try:
        with zipfile.ZipFile(zip_path, 'w') as report_zip:
            for root, dirs, files in os.walk(pdfs_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    report_zip.write(file_path, arcname=os.path.relpath(file_path, pdfs_dir))
    except Exception as e:
        current_app.logger.error(f"Failed to create zip file: {e}")
        return "Error generating reports", 500

    if not os.path.exists(zip_path):
        current_app.logger.error(f"Zip file not found at path: {zip_path}")
        return "Error generating reports", 500

    try:
        return send_file(zip_path, as_attachment=True)
    except Exception as e:
        current_app.logger.error(f"Failed to send zip file: {e}")
        return "Error sending the report", 500

  

@promote.route('/generate_excel_reports/<int:class_id>/<int:year>')
@school_it_admin_required
@login_required
def generate_excel_reports(class_id, year):
    class_ = Class.query.get_or_404(class_id)

    user_downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads", "reports")
    os.makedirs(user_downloads_dir, exist_ok=True)

    excel_dir = os.path.join(user_downloads_dir, f'class_{class_id}_year_{year}')
    os.makedirs(excel_dir, exist_ok=True)

    school = class_.company  
    school_title = school.title
    school_address = school.location

    try:
        for student in class_.students:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"{student.user.first_name} {student.user.last_name}"

            # Add school and class details
            sheet.append([school_title])
            sheet.append([school_address])
            sheet.append([])  # Empty row for spacing

            # Add student information headers
            student_info_headers = [
                _('Nom'), 
                _('Prénom'), 
                _('Genre'), 
                _('Classe'), 
                _('Année Scolaire')
            ]
            sheet.append(student_info_headers)

            # Add student information data
            student_info_data = [
                student.user.first_name,          # First Name
                student.user.last_name,           # Last Name
                student.user.gender,              # Gender
                class_.name,                      # Class Name
                year                              # Academic Year
            ]
            sheet.append(student_info_data)
            sheet.append([])  # Empty row for spacing

            # Add headers for the grade section
            grade_headers = [
                _('Session'), 
                _('Matière'), 
                _('Nom de l\'examen'), 
                _('Type d\'examen'), 
                _('Note')
            ]
            sheet.append(grade_headers)

            # Style the header row
            for cell in sheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
            
            # Adjust column widths for better readability
            column_widths = [15, 15, 20, 20, 10]
            for i, width in enumerate(column_widths):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width
            
            # Generate the token for the report
            report_token = generate_report_token(student.user_id, year)

            # Generate the QR code URL
            qr_code_url = generate_qr_code(report_token)

            # Add QR Code to the Excel with a reduced size
            qr_img = qrcode.make(qr_code_url)
            qr_img = qr_img.resize((100, 100))  # Resize QR code
            qr_img_path = os.path.join(excel_dir, f"{secure_filename(student.user.last_name)}_{secure_filename(student.user.first_name)}_qr.png")
            qr_img.save(qr_img_path)

            # Insert QR code image into the Excel sheet
            img = openpyxl.drawing.image.Image(qr_img_path)
            img.anchor = 'F5'  # Adjust as needed for placement
            sheet.add_image(img)

            # Fetch the academic year to ensure we are looking at the right sessions
            academic_year = AcademicYear.query.filter_by(id=year).first()
            if academic_year:
                sessions = Session.query.filter(Session.academic_year_id == academic_year.id).all()
            else:
                sessions = []

            for session in sessions:
                for subject in class_.subjects:
                    # Fetch the grades for this student, session, and subject
                    grades = Grade.query.filter_by(student_id=student.user_id, session_id=session.id, subject_id=subject.id).all()

                    # Add each grade with the corresponding exam details
                    for grade in grades:
                        exam = grade.exam  # Fetch the associated exam from the grade

                        if exam:
                            # Append the exam details to the sheet
                            sheet.append([
                                session.name,       # Session name
                                subject.name,       # Subject name
                                exam.name,          # Exam name
                                exam.type,          # Exam type
                                grade.value         # Grade value
                            ])

            # Save the workbook for the student
            excel_file_name = f"{secure_filename(student.user.last_name)}_{secure_filename(student.user.first_name)}_report.xlsx"
            excel_file_path = os.path.join(excel_dir, excel_file_name)
            workbook.save(excel_file_path)

        zip_filename = f'class_{class_id}_year_{year}_reports.zip'
        zip_path = os.path.join(user_downloads_dir, zip_filename)

        with zipfile.ZipFile(zip_path, 'w') as report_zip:
            for root, dirs, files in os.walk(excel_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    report_zip.write(file_path, arcname=os.path.relpath(file_path, excel_dir))

        # Return the zip file as a downloadable attachment
        return send_file(zip_path, as_attachment=True)

    except Exception as e:
        current_app.logger.error(f"Error generating Excel reports: {e}")
        return "Error generating reports", 500



@promote.route('/view_report/<string:token>')
@login_required
def view_student_report_fr(token):
    """
    View the student's report based on the token.
    
    Args:
    - token: The unique token identifying the report.

    Returns:
    - Rendered HTML report for the student.
    """
    report_token = ReportToken.query.filter_by(token=token).first()

    if not report_token:
        return "Invalid token or report not found.", 404

    student_id = report_token.student_id
    session_id = report_token.session_id

    report_data = compute_student_report_fr(student_id, session_id)
    
    return render_template('reports/academics/FR/session_report_1.html', data=report_data)

